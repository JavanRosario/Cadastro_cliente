from pathlib import Path
import openpyxl


def init_exel():
    #initializing xlsx file
    caminho = Path('Clientes.xlsx')
    if not caminho.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Nome completo'
        ws['B1'] = 'Contato'
        ws['C1'] = 'Idade'
        ws['D1'] = 'Endereço'
        ws['E1'] = 'Gênero'
        ws['F1'] = 'Observações'
        wb.save('Clientes.xlsx')


def data_save(nome, contato, idade, endereco, genero, observacoes):
    #saving data in columns, splitting by names
    wb = openpyxl.load_workbook('Clientes.xlsx')
    ws = wb.active
    linha = ws.max_row + 1
    ws.cell(row=linha, column=1, value=nome)
    ws.cell(row=linha, column=2, value=contato)
    ws.cell(row=linha, column=3, value=idade)
    ws.cell(row=linha, column=4, value=endereco)
    ws.cell(row=linha, column=5, value=genero)
    ws.cell(row=linha, column=6, value=observacoes)
    wb.save('Clientes.xlsx')


def clear_exel_data():
    #function to clear all written data
    wb = openpyxl.load_workbook('Clientes.xlsx')
    ws = wb.active
    ws.delete_rows(2, ws.max_row-1)
    wb.save('Clientes.xlsx')

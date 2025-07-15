import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import xlrd
from pathlib import Path
from openpyxl import workbook

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("dark-blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.all_system()

    def layout_config(self):
        self.title("Sistema de Cadastro de Clientes")
        self.geometry("800x500")
        self.resizable(False, False)  # disable resizing
        self.maxsize(800, 500)
        self.minsize(800, 500)

    def appearance(self):
        # Appearence mode ui
        self.appearance_mode_label = ctk.CTkLabel(
            self, text='tema', bg_color='transparent', text_color=['#000', '#fff']).place(x=50, y=430)
        self.appearance_mode_options = ctk.CTkOptionMenu(
            self, values=['Light', 'Dark', 'System'], command=self.change_appearance).place(x=50, y=460)

    def change_appearance(self, new_appearance_mode):
        ctk.set_appearance_mode(new_appearance_mode)

    def all_system(self):
        # x,y patterns
        x_label = 50
        x_entry = 220
        y_start = 120
        y_step = 50
        button_y = y_start + 4 * y_step + 20

        # header frame title
        frame = ctk.CTkFrame(self, width=800, height=50,
                             corner_radius=0, bg_color="green", fg_color="gray")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text='Sistema Cadastro de Clientes', font=(
            "Segoe UI", 25), text_color="#fff")
        title.place(relx=0.5, rely=0.5, anchor='center')

        small_box = ctk.CTkLabel(self, text='Prencha todos os campos:', font=(
            "Segoe UI", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        file = Path('Clientes.xlsx')
        if file.exists():
            pass
        else:
            file= Workbook()
            file = file.active
            sheet['A1'] = 'Nome completo'
            sheet['B2'] = 'Contato'
            sheet['C3'] = 'Idade'
            sheet['D4'] = 'Gênero'
            sheet['E1'] = 'Endereço'
            sheet['F1'] = 'Observações'

            file.save('Clientes.xlsx')

        def submit():
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            adress = address_value.get()
            gender = gender_combobox.get()
            obs = obs_entry.get(0.0, END)

            file = openpyxl.load_workbook('Clientes.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row+1, value=name)
            sheet.cell(column=2, row=sheet.max_row, value=contact)
            sheet.cell(column=3, row=sheet.max_row, value=age)
            sheet.cell(column=4, row=sheet.max_row, value=adress)
            sheet.cell(column=5, row=sheet.max_row, value=gender)
            sheet.cell(column=6, row=sheet.max_row, value=obs)

            file.save(r'Clientes.xlsx')
            messagebox.showinfo('Systema', 'Dados santos com sucesso!')

        def clear():
            name_value.set('')
            contact_value.set('')
            age_value.set('')
            address_value.set('')
            obs_entry.delete("0.0", END)

        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()

        # name
        input_name = ctk.CTkLabel(self, text='Nome Completo:', font=(
            "Segoe UI", 16), text_color=["#000", "#fff"])
        input_name.place(x=x_label, y=y_start)
        name_entry = ctk.CTkEntry(self, width=500, textvariable=name_value, font=(
            "Segoe UI", 16), fg_color='transparent')
        name_entry.place(x=x_entry, y=y_start-5)
        # contact
        input_contact = ctk.CTkLabel(self, text='Número de contato:', font=(
            "Segoe UI", 16), text_color=["#000", "#fff"])
        input_contact.place(x=x_label, y=y_start + y_step)
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=(
            "Segoe UI", 16), fg_color='transparent')
        contact_entry.place(x=x_entry, y=y_start + y_step - 5)
        # age
        input_age = ctk.CTkLabel(self, text='Idade:', font=(
            "Segoe UI", 16), text_color=["#000", "#fff"])
        input_age.place(x=440, y=y_start + y_step)
        age_entry = ctk.CTkEntry(self, width=80,  textvariable=age_value, font=(
            "Segoe UI", 16), fg_color='transparent')
        age_entry.place(x=500, y=y_start + y_step - 5)
        # gender
        input_gender = ctk.CTkLabel(self, text='Gênero:', font=(
            "Segoe UI", 16), text_color=["#000", "#fff"])
        input_gender.place(x=600, y=y_start + y_step)
        gender_combobox = ctk.CTkComboBox(
            self, values=['Masculino', 'Feminino'], font=('Segoe UI', 15), width=120,)
        gender_combobox.set('Masculino')
        gender_combobox.place(x=670, y=y_start + y_step - 5)
        # address
        input_address = ctk.CTkLabel(self, text='Endereço:', font=(
            "Segoe UI", 16), text_color=["#000", "#fff"])
        input_address.place(x=x_label, y=y_start + 2 * y_step)
        address_entry = ctk.CTkEntry(self, width=500, textvariable=address_value, font=(
            "Segoe UI", 16), fg_color='transparent')
        address_entry.place(x=x_entry, y=y_start + 2 * y_step - 5)
        # OBS
        input_obs = ctk.CTkLabel(self, text='Observações:', font=(
            "Segoe UI", 16), text_color=["#000", "#fff"])
        input_obs.place(x=x_label, y=y_start + 3 * y_step + 4)
        obs_entry = ctk.CTkTextbox(self, width=670, height=150, font=(
            'Arial', 18), bg_color='#aaa', border_width=2, fg_color='transparent')
        obs_entry.place(x=x_label, y=y_start + 3 * y_step + 30)

        button_submit = ctk.CTkButton(
            self, text='Salvar Dados', command=submit, fg_color='#151', hover_color='#131')
        button_submit.place(x=220, y=button_y-82)
        button_clear = ctk.CTkButton(
            self, text='Limpar Campos', command=clear, fg_color='#555', hover_color='#333')
        button_clear.place(x=400, y=button_y-82)


if __name__ == '__main__':
    app = App()
    app.mainloop()
